import customtkinter as ctk
from tkinter import *
from tkinter import messagebox
import openpyxl, xlrd
import pathlib
from openpyxl import Workbook

# fazendo a aparencia padrão do sistema
ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")


class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.layout_config()
        self.appearence()
        self.todo_sistema()

    def layout_config(self):
        self.title("Sistema de Gestão de Clientes")
        self.geometry('700x500')

    def appearence(self):#aparencia do sistema
        self.lb_apm = ctk.CTkLabel(self, text="Tema", bg_color="transparent", text_color=['#000', "#fff"]).place(x=50, y=430)
        self.opt_apm = ctk.CTkOptionMenu(self, values=["Light", "Dark", "System"], command=self.change_apm).place(x=50, y=460)

    def todo_sistema(self):#aqui ficará
        frame = ctk.CTkFrame(self, width=700, height=50, corner_radius=0, bg_color='teal', fg_color='teal')#faixa azul
        frame.place(x=0, y=10)
        title = ctk.CTkLabel(frame, text='Sistema de Gestão de Clientes',font=("Century Gothic bold", 24), text_color="#fff")

        span = ctk.CTkLabel(self, text='Por favor preencha os campos do formulário',font=("Century Gothic bold", 16), text_color=["#000", "#fff"]).place(x=50, y=70)

        ficheiro = pathlib.Path("Cliente.xlsx")

        if ficheiro.exists():
            pass
        else:
            ficheiro = Workbook()
            folha = ficheiro.active
            folha['A1'] = "Nome completo"
            folha['B1'] = "Contato"
            folha['C1'] = "Idade"
            folha['D1'] = "Gênero"
            folha['E1'] = "Endereço"
            folha['F1'] = "Observações"

            ficheiro.save("Cliente.xlsx")
        # funções dos botões
        def submit():

            #pegando os dados
            name = name_value.get()
            contacto = contato_value.get()
            age = idade_value.get()
            gender = gender_combobox.get()
            obs = obs_entry.get(0.0, END)
            address = endereco_value.get()

            if (name=='' or contacto=='' or age=='' or address==''):
                messagebox.showerror("Systma", "Erro!\nPor Favor preencha todos os campos.")
            #elif name==name:
                #messagebox.showerror("System", "Erro!\nUsuario ja Cadastrado")


            else:

                ficheiro = openpyxl.load_workbook('Cliente.xlsx')
                folha = ficheiro.active
                folha.cell(column=1, row=folha.max_row + 1, value=name)
                folha.cell(column=2, row=folha.max_row, value=contacto)
                folha.cell(column=3, row=folha.max_row, value=age)
                folha.cell(column=4, row=folha.max_row, value=gender)
                folha.cell(column=5, row=folha.max_row, value=address)
                folha.cell(column=6, row=folha.max_row, value=obs)

                ficheiro.save(r"Cliente.xlsx")
                messagebox.showinfo("Systema", "Dados salvos com sucesso!")

        def clear():
            name_value.set("")
            contato_value.set("")
            idade_value.set("")
            obs_entry.delete(0.0, END)
            endereco_value.set("")

        #Variaveis de texto
        name_value = StringVar()
        contato_value = StringVar()
        idade_value = StringVar()
        endereco_value = StringVar()


        #Entrada onde o cliente ira digitar
        name_entry = ctk.CTkEntry(self, width=350, textvariable=name_value, font=("Century Gothic Bold", 16), fg_color="transparent")
        contato_entry = ctk.CTkEntry(self, width=200, textvariable=contato_value, font=("Century Gothic Bold", 16), fg_color="transparent")
        idade_entry = ctk.CTkEntry(self, width=150, textvariable=idade_value, font=("Century Gothic Bold", 16), fg_color="transparent")
        endereco_entry = ctk.CTkEntry(self, width=200, textvariable=endereco_value, font=("Century Gothic Bold", 16), fg_color="transparent")

       #Combobox para o genero do cliente
        gender_combobox = ctk.CTkComboBox(self, values=["Masculino", "Feminino"], font=("Century Gothic Bold", 14))
        gender_combobox.set("Masculino")

        #Entrada de observações
        obs_entry = ctk.CTkTextbox(self, width=500, height=150, font=("arial", 18), border_color="#aaa", border_width=2, fg_color="transparent")

        #labels, onde fica o subtitulo
        lb_nome = ctk.CTkLabel(self, text='Nome', font=("Century Gothic bold", 16),text_color=["#000", "#fff"])
        lb_contato = ctk.CTkLabel(self, text='Contato', font=("Century Gothic bold", 16),text_color=["#000", "#fff"])
        lb_idade = ctk.CTkLabel(self, text='Idade', font=("Century Gothic bold", 16),text_color=["#000", "#fff"])
        lb_genero = ctk.CTkLabel(self, text='Genero', font=("Century Gothic bold", 16),text_color=["#000", "#fff"])
        lb_endereco = ctk.CTkLabel(self, text='Endereço', font=("Century Gothic bold", 16),text_color=["#000", "#fff"])
        lb_obs = ctk.CTkLabel(self, text='Observações:', font=("Century Gothic bold", 16),text_color=["#000", "#fff"])

        #botão para salvar os dados
        bnt_submit = ctk.CTkButton(self, text= "Salvar dados".upper(), command=submit, fg_color="#151", hover_color="#131").place(x=300, y=420)
        bnt_clear = ctk.CTkButton(self, text="Limpar dados".upper(), command=clear, fg_color="#555",hover_color="#333").place(x=500, y=420)


        #Posicionando os elementos na tela

        title.place(x=200, y=10)

        lb_nome.place(x=50, y=120)
        name_entry.place(x=50, y=150)

        lb_contato.place(x=450, y=120)
        contato_entry.place(x=450, y=150)

        lb_idade.place(x=300, y=190)
        idade_entry.place(x=300, y=220)

        lb_endereco.place(x=50, y=190)
        endereco_entry.place(x=50, y=220)

        lb_genero.place(x=500, y=190)
        gender_combobox.place(x=500, y=220)

        lb_obs.place(x=50, y=260)
        obs_entry.place(x=150, y=260)

    def change_apm(self, nova_aparencia):
        ctk.set_appearance_mode(nova_aparencia)


if __name__ == '__main__':
    app = App()
    app.mainloop()
